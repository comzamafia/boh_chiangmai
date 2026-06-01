"""Build a formatted Excel export of all production recipes."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

headers = [
    "ID", "Recipe Name", "Category", "Yield Amount", "Yield Unit",
    "Prep Time (min)", "Cook Time (min)", "Labor Cost/hr (฿)",
    "Energy Cost/batch (฿)", "Selling Price (฿)", "Delivery Price (฿)",
    "Is Main Sauce", "Is Sub Recipe", "Linked Ingredient ID", "Image URL",
    "# Ingredients", "Ingredient Cost (฿)", "Labor Cost (฿)",
    "Energy Cost (฿)", "Total Cost/batch (฿)",
    "Food Cost % (Dine-in)", "Food Cost % (Delivery)",
    "Ingredient Names", "Instructions",
    "Created At (ICT)", "Updated At (ICT)",
]

data = [
    ["R006","Fresh Spring Rolls","Appetizers",4,"piece",20,0,50,0.5,"","","No","No","","",0,0,16.67,0.5,17.17,"N/A","N/A","","","21/2/2569 18:58:50","21/2/2569 18:58:50"],
    ["R005","Chili Oil (Nam Prik Pao)","Main Sauce",5,"L",30,45,50,12,"","","Yes","No","","",5,238.5,62.5,12,313,"N/A","N/A","Chili Flakes; Dried Shrimp; Fish Sauce (Nam Pla); Palm Sugar; Vegetable Oil","1. Toast dried chilies and garlic separately. 2. Blend coarsely. 3. Fry in vegetable oil on medium heat. 4. Add dried shrimp, palm sugar, fish sauce. 5. Simmer 20 min. Cool and store.","21/2/2569 18:58:50","21/2/2569 18:58:50"],
    ["R003","Pad Thai Sauce Base","Main Sauce",10,"L",15,60,50,15,"","","Yes","No","","",3,295,62.5,15,372.5,"N/A","N/A","Fish Sauce (Nam Pla); Palm Sugar; Tamarind Paste","1. Soak tamarind in warm water. Strain through sieve. 2. Combine tamarind water, palm sugar, fish sauce. 3. Simmer until sugar dissolves. 4. Reduce 20%. 5. Cool. Store refrigerated.","21/2/2569 18:58:50","21/2/2569 18:58:50"],
    ["R007","Tom Yum Broth Base","Main Sauce",8,"L",10,40,50,10,"","","Yes","No","","",0,0,41.67,10,51.67,"N/A","N/A","","","21/2/2569 18:58:50","21/2/2569 18:58:50"],
    ["R004","Chicken Pad Thai","Noodles",1,"serving",8,7,50,2,"","","No","No","","",7,36.97,12.5,2,51.47,"N/A","N/A","Bean Sprouts; Chicken Breast; Egg (Chicken); Pad Thai Noodles (Sen Lek); Pad Thai Sauce Base; Soy Sauce (Light); Vegetable Oil","1. Marinate chicken. 2. Soak noodles. 3. Cook chicken through. 4. Add noodles, sauce and egg. 5. Fold in vegetables and serve.","21/2/2569 18:58:50","21/2/2569 18:58:50"],
    ["R001","Classic Pad Thai","Noodles",1,"serving",5,5,50,2,"","","No","No","","",8,31.07,8.33,2,41.4,"N/A","N/A","Bean Sprouts; Egg (Chicken); Pad Thai Noodles (Sen Lek); Pad Thai Sauce Base; Peanuts (Raw); Spring Onion; Tofu (Firm); Vegetable Oil","1. Soak noodles 30 min. 2. Heat wok. 3. Stir-fry tofu. 4. Crack egg. 5. Add noodles and sauce. 6. Toss with bean sprouts. 7. Serve with peanuts.","21/2/2569 18:58:50","21/2/2569 18:58:50"],
    ["R008","Pad See Ew","Noodles",1,"serving",5,5,50,2,"","","No","No","","",0,0,8.33,2,10.33,"N/A","N/A","","","21/2/2569 18:58:50","21/2/2569 18:58:50"],
    ["R002","Shrimp Pad Thai","Noodles",1,"serving",5,7,50,2.5,"","","No","No","","",8,90.52,10,2.5,103.02,"N/A","N/A","Bean Sprouts; Dried Shrimp; Egg (Chicken); Pad Thai Noodles (Sen Lek); Pad Thai Sauce Base; Peanuts (Raw); Tiger Shrimp; Vegetable Oil","1. Soak noodles 30 min. 2. Stir-fry shrimp. 3. Add noodles and sauce. 4. Fold in egg, bean sprouts. 5. Garnish with dried shrimp and peanuts.","21/2/2569 18:58:50","21/2/2569 18:58:50"],
]

AMBER       = "F59E0B"
WHITE       = "FFFFFF"
BORDER_CLR  = "E5E7EB"

def thin(color=BORDER_CLR):
    s = Side(border_style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

wb = Workbook()

# ── Sheet 1: Recipes ──────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Recipes"
ws.sheet_view.showGridLines = False

# Header
for ci, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=ci, value=h)
    c.font       = Font(name="Arial", bold=True, color=WHITE, size=10)
    c.fill       = PatternFill("solid", fgColor=AMBER)
    c.alignment  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border     = thin()
ws.row_dimensions[1].height = 38

# Column widths
widths = [
    8, 28, 14, 11, 9, 13, 12, 15, 16, 13, 14,
    11, 11, 18, 22, 12, 16, 13, 13, 17, 18, 18,
    45, 52, 20, 20,
]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Numeric column indices (1-based)
num_cols = {4, 6, 7, 8, 9, 10, 11, 16, 17, 18, 19, 20}

cat_bg = {
    "Appetizers": ("EFF6FF", "DBEAFE"),
    "Main Sauce":  ("FFF8E7", "FDE68A"),
    "Noodles":     ("F0FDF4", "D1FAE5"),
}

for ri, row in enumerate(data, 2):
    cat  = row[2]
    even = ri % 2 == 0
    bg   = cat_bg.get(cat, (WHITE, "F3F4F6"))[0 if even else 1]
    is_main = row[11] == "Yes"

    for ci, val in enumerate(row, 1):
        c = ws.cell(row=ri, column=ci, value=val if val != "" else None)
        c.font      = Font(name="Arial", bold=(is_main and ci == 2), size=10)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(
            horizontal="right" if ci in num_cols else "left",
            vertical="center",
            wrap_text=True,
        )
        c.border = thin()
        if ci in num_cols and isinstance(val, float):
            c.number_format = "#,##0.00"
        elif ci in num_cols and isinstance(val, int) and ci != 1:
            c.number_format = "#,##0"
    ws.row_dimensions[ri].height = 56

ws.freeze_panes = "A2"
ws.auto_filter.ref = "A1:" + get_column_letter(len(headers)) + "1"

# ── Sheet 2: Summary ──────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Summary")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 14
ws2.column_dimensions["C"].width = 18
ws2.column_dimensions["D"].width = 60

# Title
ws2.merge_cells("A1:D1")
t = ws2["A1"]
t.value     = "Chiang Mai BOH — Recipe Export Summary"
t.font      = Font(name="Arial", bold=True, size=18, color=WHITE)
t.fill      = PatternFill("solid", fgColor=AMBER)
t.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 48

ws2.merge_cells("A2:D2")
sub = ws2["A2"]
sub.value     = "Exported from production database · boh-chiangmai.vercel.app · 20 May 2026"
sub.font      = Font(name="Arial", italic=True, size=10, color="6B7280")
sub.alignment = Alignment(horizontal="center")
ws2.row_dimensions[2].height = 22

# Stats row
ws2.row_dimensions[4].height = 20
ws2.row_dimensions[5].height = 52
ws2.row_dimensions[6].height = 16

stats = [
    ("Total Recipes", 8,  "10B981"),
    ("Categories",    3,  "3B82F6"),
    ("Main Sauces",   3,  "F59E0B"),
    ("Noodle Dishes", 4,  "8B5CF6"),
]
for i, (label, val, clr) in enumerate(stats):
    col = i + 1
    lc = ws2.cell(row=4, column=col, value=label)
    lc.font      = Font(name="Arial", size=9, color="6B7280")
    lc.alignment = Alignment(horizontal="center")

    vc = ws2.cell(row=5, column=col, value=val)
    vc.font      = Font(name="Arial", bold=True, size=28, color=clr)
    vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.fill      = PatternFill("solid", fgColor="F9FAFB")
    b = Side(border_style="medium", color=clr)
    vc.border    = Border(bottom=b)

# Category table
cat_headers = ["Category", "Count", "Has Main Sauce", "Recipes"]
for ci, ch in enumerate(cat_headers, 1):
    c = ws2.cell(row=7, column=ci, value=ch)
    c.font      = Font(name="Arial", bold=True, color=WHITE, size=10)
    c.fill      = PatternFill("solid", fgColor="374151")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = thin("374151")
ws2.row_dimensions[7].height = 26

cat_rows = [
    ("Appetizers", 1, "No",  "Fresh Spring Rolls"),
    ("Main Sauce",  3, "Yes", "Chili Oil (Nam Prik Pao); Pad Thai Sauce Base; Tom Yum Broth Base"),
    ("Noodles",     4, "No",  "Chicken Pad Thai; Classic Pad Thai; Pad See Ew; Shrimp Pad Thai"),
]
cat_bgs2 = ["DBEAFE", "FDE68A", "D1FAE5"]
for ri, (cat, cnt, ms, names) in enumerate(cat_rows, 8):
    bg = cat_bgs2[ri - 8]
    row_vals = [cat, cnt, ms, names]
    for ci, v in enumerate(row_vals, 1):
        c = ws2.cell(row=ri, column=ci, value=v)
        c.font      = Font(name="Arial", size=10)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center" if ci == 2 else "left", vertical="center", wrap_text=True)
        c.border    = thin()
    ws2.row_dimensions[ri].height = 36

out_path = r"C:\PADTHAI-CHAIYO-BOH\backofhouse\scripts\recipes-export.xlsx"
wb.save(out_path)
print("Saved:", out_path)
